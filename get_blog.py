import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

blog_wb=load_workbook("search_keywords.xlsx")
blog_ws=blog_wb.active
#blog_wb.get_sheet_by_name 이름으로 불러오기
#blog_wb.worksheets[0] 시트로 불러오기
#rows=blog_ws.rows
rows=list(blog_ws.rows)[1:] #list로 형변환 후 ,맨위줄 빼고 2번째줄부터 끝까지

#전체 데이터를 모두 출력하기
for row in rows:
    #print(row[0].value) 한줄에 있는 엑셀파일 출력하기
    for column in row:
        print(column.value)

#exit()

        keyword=column.value
        url="https://search.naver.com/search.naver?where=post&sm=tab_jum&query="+keyword
        data=requests.get(url)

        if data.status_code != requests.codes.ok:
            print("접속실패")
            exit()

        wb=Workbook()
        ws=wb.active
        ws.cell(1,1,"키워드")
        ws.cell(1,2,"제목")
        ws.cell(1,3,"링크")
        ws.cell(1,4,"요약")

        html=BeautifulSoup(data.text,"html.parser")
        blog_data=html.select(".type01 > li")

        #for blogs in blog_data:
        for index, blogs in enumerate(blog_data,start=2):
            title=blogs.select_one(".sh_blog_title")
            link=title.attrs['href']
            description=blogs.select('dd')[1].text
            #print(title.text,link,description)

            ws.cell(index,1,keyword)
            ws.cell(index,2,title.text)
            ws.cell(index,3,link)
            ws.cell(index,4,description)

    wb.save(f"blogs_{keyword}.xlsx")